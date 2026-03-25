from openpyxl import load_workbook

from util.constant import (
    DO_ANALYSIS_TABLE,
    EXCEL_FILE,
    ANALYSIS_SHEET_NAME,
    TO_BE_CHECKED_TABLE_SHEET_NAME,
    ANALYSIS_TABLE,
    ANALYSIS_COLUMNS,
    DO_ANALYSIS_START_COL,
    IST_TIMEZONE,

)

from util.excel_helpers import (
    get_table,
    get_table_rows,
    get_table_column_index_map,
    insert_row_into_table,
    get_ist_datetime
)
import pytz
from datetime import datetime


# @app.get("/analysis")
# --------------------------------------------------
# Scoring APIs
# --------------------------------------------------
def get_analysis_service(date = None):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[ANALYSIS_SHEET_NAME]
    ws_to_check = wb[TO_BE_CHECKED_TABLE_SHEET_NAME]

    # -------------------------------
    # PART 1: Analysis Checks (Table6 / DO_ANALYSIS_TABLE)
    # -------------------------------
    checks_table = ws_to_check.tables[DO_ANALYSIS_TABLE]
    start_cell, end_cell = checks_table.ref.split(":")
    start_row = ws_to_check[start_cell].row + 1  # skip header
    end_row = ws_to_check[end_cell].row

    analysis_checks = []

    for row in range(start_row, end_row + 1):
        check_id = ws_to_check.cell(row=row, column=DO_ANALYSIS_START_COL).value
        entry = ws_to_check.cell(row=row, column=DO_ANALYSIS_START_COL + 1).value

        if check_id is None and entry is None:
            continue

        analysis_checks.append({
            "id": check_id,
            "entry": entry
        })

    # -------------------------------
    # PART 2: Analysis Entries (ANALYSIS_TABLE)
    # -------------------------------
    table = get_table(ws, ANALYSIS_TABLE)
    col_map = get_table_column_index_map(ws, table)
    rows = get_table_rows(ws, table, col_map)

    analysis_entries = []

    for row in rows:
        cell_value = row[ANALYSIS_COLUMNS["datetime"]]
        if not cell_value:
            continue

        analysis_entries.append({
            "entry_id": row[ANALYSIS_COLUMNS["entry_id"]],
            "analysis": row[ANALYSIS_COLUMNS["analysis"]],
            "improvement": row[ANALYSIS_COLUMNS["improvement"]],
            "status": row[ANALYSIS_COLUMNS["status"]],
            "datetime": row[ANALYSIS_COLUMNS["datetime"]]
        })

    return {
        "date": date,
        "analysis_checks": analysis_checks,
        "analysis_entries": analysis_entries
    }

# @app.post("/save-analysis")
# -------------------------------
# Save Analysis 
# -------------------------------
def save_analysis_service(payload):

    wb = load_workbook(EXCEL_FILE)
    ws = wb[ANALYSIS_SHEET_NAME]

    table = get_table(ws, ANALYSIS_TABLE)
    col_map = get_table_column_index_map(ws, table)
    rows = get_table_rows(ws, table, col_map)

    ist_datetime = get_ist_datetime()
    today = ist_datetime[:10]

    matched = False

    # -------- UPDATE if entry_id exists for today --------
    for row in rows:
        cell_value = row[ANALYSIS_COLUMNS["datetime"]]
        if not cell_value:
            continue

        row_date = str(cell_value)[:10]

        if row_date == today and row[ANALYSIS_COLUMNS["entry_id"]] == payload.entry_id:
            ws.cell(
                row=row["_row"],
                column=col_map[ANALYSIS_COLUMNS["datetime"]],
                value=ist_datetime
            )
            ws.cell(
                row=row["_row"],
                column=col_map[ANALYSIS_COLUMNS["analysis"]],
                value=payload.analysis
            )
            ws.cell(
                row=row["_row"],
                column=col_map[ANALYSIS_COLUMNS["improvement"]],
                value=payload.improvement
            )
            ws.cell(
                row=row["_row"],
                column=col_map[ANALYSIS_COLUMNS["status"]],
                value=payload.status
            )
            matched = True
            break

    # -------- INSERT if not found --------
    if not matched:
        insert_row_into_table(
            ws,
            table,
            col_map,
            {
                ANALYSIS_COLUMNS["datetime"]: ist_datetime,
                ANALYSIS_COLUMNS["entry_id"]: payload.entry_id,
                ANALYSIS_COLUMNS["analysis"]: payload.analysis,
                ANALYSIS_COLUMNS["improvement"]: payload.improvement,
                ANALYSIS_COLUMNS["status"]: payload.status
            }
        )

    wb.save(EXCEL_FILE)

    return {
        "message": "Analysis saved successfully",
        "entry_id": payload.entry_id,
        "updated": matched,
        "saved_at_ist": ist_datetime
    }



# @app.post("/analysis-checks-save")
# -------------------------------
# Check For analysis Entry ID 
# -------------------------------
def save_analysis_checks_service(items):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[ANALYSIS_SHEET_NAME]
    ws_to_check = wb[TO_BE_CHECKED_TABLE_SHEET_NAME]

    # Find Table7
    table = ws_to_check.tables[DO_ANALYSIS_TABLE]

    # Get last row of the table
    start_row = ws_to_check[table.ref.split(":")[1]].row + 1

    for item in items:
        ws_to_check.cell(row=start_row, column=DO_ANALYSIS_START_COL).value = item.id
        ws_to_check.cell(row=start_row, column=DO_ANALYSIS_START_COL + 1).value = item.entry
        start_row += 1

    # Extend table range
    end_col_letter = table.ref.split(":")[1][0]
    table.ref = f"{table.ref.split(':')[0]}:{end_col_letter}{start_row - 1}"

    wb.save(EXCEL_FILE)

    return {
        "status": "success",
        "rows_added": len(items)
    }

# @app.post("/sync-analysis")
# -------------------------------
# Sync Analysis 
# -------------------------------
def sync_analysis_service(entries):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[ANALYSIS_SHEET_NAME]

    table = ws.tables[ANALYSIS_TABLE]

    # Table range
    start_row = ws[table.ref][0][0].row + 1
    end_row = ws[table.ref][-1][0].row

    # Current IST time
    ist = pytz.timezone(IST_TIMEZONE)
    now_ist = datetime.now(ist).strftime("%Y-%m-%d %H:%M:%S")

    # Header mapping
    header_row = ws[start_row - 1]
    col_index = {cell.value: idx + 1 for idx, cell in enumerate(header_row)}

    entry_id_col = col_index[ANALYSIS_COLUMNS["entry_id"]]

    # Input entry_ids (deduplicated)
    input_map = {
        item["entry_id"]: item
        for item in entries
        if "entry_id" in item
    }
    input_ids = set(input_map.keys())

    rows_to_delete = []

    # Iterate existing table rows
    for row in range(start_row, end_row + 1):
        excel_entry_id = ws.cell(row=row, column=entry_id_col).value

        # DELETE if not in input
        if excel_entry_id not in input_ids:
            rows_to_delete.append(row)
            continue

        # UPDATE if in input
        payload = input_map[excel_entry_id]

        # Update datetime
        ws.cell(
            row=row,
            column=col_index[ANALYSIS_COLUMNS["datetime"]],
            value=now_ist
        )

        # Update provided fields only
        for key, excel_col in ANALYSIS_COLUMNS.items():
            if key in payload and key != "datetime":
                ws.cell(
                    row=row,
                    column=col_index[excel_col],
                    value=payload[key]
                )

    # Delete rows bottom-up
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)

    wb.save(EXCEL_FILE)

    return {
        "message": "Analysis table synced successfully",
        "updated": len(input_ids),
        "deleted": len(rows_to_delete)
    }
