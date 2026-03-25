from openpyxl import load_workbook
import pytz

from util.constant import (
    EXCEL_FILE,
    GOAL_SHEET_NAME,
    GOAL_COLUMNS,
    GOAL_TABLE
)

from util.excel_helpers import (
    get_goal_rows,
    get_table,
    get_table_column_index_map,
    get_table_rows,
    insert_row_into_table,
    get_ist_datetime
)
from datetime import datetime

# @app.get("/goals")
# --------------------------------------------------
# Gaol APIs
# --------------------------------------------------
def get_goals_service(date: str):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[GOAL_SHEET_NAME]

    table, col_map, rows = get_goal_rows(ws)

    goals = []

    for row in rows:
        cell_value = row[GOAL_COLUMNS["datetime"]]
        if not cell_value:
            continue

        if date is None or str(cell_value)[:10] == date:
            last_date_raw = row[GOAL_COLUMNS["last_date"]]
            goals.append({
                "goal_id": row[GOAL_COLUMNS["goal_id"]],
                "goal": row[GOAL_COLUMNS["goal"]],
                "goal_type": row[GOAL_COLUMNS["goal_type"]],
                "last_date": str(last_date_raw)[:10] if last_date_raw else "",
                "status": row[GOAL_COLUMNS["status"]],
            })

    return {
        "date": date,
        "goals": goals
    }


# @app.post("/save-goals")
def save_goals_service(payload):
    """
    This service synchronizes goals from the frontend with the Excel goal tracker
    for the current date.

    It performs the following operations:

    1. Loads the Excel workbook and accesses the goal sheet.
    2. Reads existing goal records from the Excel table.
    3. Identifies all incoming goal IDs from the request payload.
    4. Deletes goals from Excel that:
    - Belong to the current date, and
    - Are not present in the incoming request.
    - Entire row is removed and remaining rows shift upward (no gaps).
    5. Reloads the updated rows after deletion to maintain correct row indexing.
    6. Builds a lookup map of existing goals for efficient matching.
    7. Updates existing goals only if any field has changed:
    - Goal name
    - Goal type
    - Last date to achieve the goal
    - Status
    - If no changes are detected, the update is skipped.
    8. Inserts new goals that do not exist in the Excel sheet:
    - Adds them at the next available row (no empty rows in between).
    - Uses current IST datetime for the entry.
    9. Saves all changes back to the Excel file.

    Parameters:
    - payload:
        - goals: List of goal objects containing:
            - goal_id: Unique identifier of the goal
            - goal: Goal description
            - goal_type: Priority/type of goal (e.g., high, medium, low)
            - last_date: Deadline to achieve the goal
            - status: Current status (e.g., progress, done, yet)

    Returns:
    - dict:
        - message: Confirmation message indicating successful operation
        - count: Total number of goals processed from the request
    """

    wb = load_workbook(EXCEL_FILE)
    ws = wb[GOAL_SHEET_NAME]

    table, col_map, rows = get_goal_rows(ws)

    ist_datetime = get_ist_datetime()
    today = str(ist_datetime)[:10]

    incoming_ids = {g.goal_id for g in payload.goals}

    # ---------------- DELETE REMOVED GOALS ----------------
    for row in reversed(rows):
        row_date = str(row[GOAL_COLUMNS["datetime"]])[:10]

        if (
            row_date == today
            and row[GOAL_COLUMNS["goal_id"]] not in incoming_ids
        ):
            ws.delete_rows(row["_row"])  # ✅ shifts rows upward

    # ✅ IMPORTANT: reload rows after deletion
    rows = get_goal_rows(ws)[2]

    # ---------------- BUILD FAST LOOKUP ----------------
    existing_map = {}
    for row in rows:
        row_date = str(row[GOAL_COLUMNS["datetime"]])[:10]

        if row_date == today:
            existing_map[row[GOAL_COLUMNS["goal_id"]]] = row

    # ---------------- INSERT / UPDATE ----------------
    for goal in payload.goals:

        row = existing_map.get(goal.goal_id)

        # ---------------- UPDATE ----------------
        if row:
            existing_goal = row[GOAL_COLUMNS["goal"]]
            existing_type = row[GOAL_COLUMNS["goal_type"]]
            existing_last_date = row[GOAL_COLUMNS["last_date"]]
            existing_status = row[GOAL_COLUMNS["status"]]

            # ✅ ONLY update if ANY field changed
            if (
                existing_goal != goal.goal
                or existing_type != goal.goal_type
                or str(existing_last_date) != str(goal.last_date)
                or existing_status != goal.status
            ):
                ws.cell(
                    row=row["_row"],
                    column=col_map[GOAL_COLUMNS["datetime"]],
                    value=ist_datetime
                )
                ws.cell(
                    row=row["_row"],
                    column=col_map[GOAL_COLUMNS["goal"]],
                    value=goal.goal
                )
                ws.cell(
                    row=row["_row"],
                    column=col_map[GOAL_COLUMNS["goal_type"]],
                    value=goal.goal_type
                )
                ws.cell(
                    row=row["_row"],
                    column=col_map[GOAL_COLUMNS["last_date"]],
                    value=goal.last_date
                )
                ws.cell(
                    row=row["_row"],
                    column=col_map[GOAL_COLUMNS["status"]],
                    value=goal.status
                )

            # ❌ else → skip (no change)

        # ---------------- INSERT (NO GAP GUARANTEED) ----------------
        else:
            # ✅ find last non-empty row
            last_row = 0
            for r in range(ws.max_row, 0, -1):
                if any(ws.cell(row=r, column=c).value for c in col_map.values()):
                    last_row = r
                    break

            insert_row = last_row + 1  # ✅ no gaps

            ws.cell(row=insert_row, column=col_map[GOAL_COLUMNS["datetime"]], value=ist_datetime)
            ws.cell(row=insert_row, column=col_map[GOAL_COLUMNS["goal_id"]], value=goal.goal_id)
            ws.cell(row=insert_row, column=col_map[GOAL_COLUMNS["goal"]], value=goal.goal)
            ws.cell(row=insert_row, column=col_map[GOAL_COLUMNS["goal_type"]], value=goal.goal_type)
            ws.cell(row=insert_row, column=col_map[GOAL_COLUMNS["last_date"]], value=goal.last_date)
            ws.cell(row=insert_row, column=col_map[GOAL_COLUMNS["status"]], value=goal.status)

    wb.save(EXCEL_FILE)

    return {
        "message": "Goals saved successfully",
        "count": len(payload.goals)
    }


# @app.get("/goals-by-month-range")
# -----------------------------------------------------------
# Monthly Report APIs
# -----------------------------------------------------------
def get_goals_by_month_range_service(start_date, end_date, year, month, IST):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[GOAL_SHEET_NAME]

    table = get_table(ws, GOAL_TABLE)
    col_map = get_table_column_index_map(ws, table)
    rows = get_table_rows(ws, table, col_map)

    result = []

    for row in rows:
        cell_value = row[GOAL_COLUMNS["datetime"]]
        if not cell_value:
            continue

        # --- NORMALIZE DATETIME ---
        if isinstance(cell_value, datetime):
            row_dt = cell_value

        elif isinstance(cell_value, str):
            try:
                # Adjust format if needed
                row_dt = datetime.strptime(cell_value.strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                # Skip invalid date formats
                continue
        else:
            continue

    
        if row_dt.tzinfo is None:
            row_dt = IST.localize(row_dt)

        if start_date <= row_dt <= end_date:
            result.append({
                "datetime": row_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "goal_id": row[GOAL_COLUMNS["goal_id"]],
                "goal": row[GOAL_COLUMNS["goal"]],
                "goal_type": row[GOAL_COLUMNS["goal_type"]],
                "last_date": row[GOAL_COLUMNS["last_date"]],
                "status": row[GOAL_COLUMNS["status"]],
            })

    return {
        "selected_month": f"{year}-{month:02d}",
        "range": {
            "from": start_date.strftime("%Y-%m-%d"),
            "to": end_date.strftime("%Y-%m-%d")
        },
        "count": len(result),
        "goals": result
    }
