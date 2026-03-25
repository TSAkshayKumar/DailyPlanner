# Handles:
#     - daily tasks
#     - priority tasks
#     - reminders
#     - tracker
#     - score

from openpyxl import load_workbook
from collections import defaultdict

from tomlkit import table, ws
from util.constant import (
    EXCEL_FILE,
    DAILY_TASK_SHEET_NAME,
    PRIORITY_TABLE,
    REMINDER_TABLE,
    REMINDER_TABLE_SHEET_NAME,
    SCORE_TABLE,
    SCORE_TABLE_SHEET_NAME,
    TRACKER_TABLE_SHEET_NAME,
    PRIORITY_COLUMNS,
    REMINDER_COLUMNS,
    SCORE_COLUMNS,
    TRACKER_COLUMNS
)
from openpyxl.utils import range_boundaries
from util.excel_helpers import (
    get_table,
    get_table_rows,
    get_table_column_index_map,
    insert_row_into_table,
    upsert_score_for_today,
    get_tracker_column_index_map,
    find_tracker_row,
    get_ist_datetime
)

# ---------------------------------------------------  
# DAILY TASKS Get API  
# ---------------------------------------------------    
def get_daily_tasks_service(date: str):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[DAILY_TASK_SHEET_NAME]
    ws_reminder = wb[REMINDER_TABLE_SHEET_NAME]
    ws_score = wb[SCORE_TABLE_SHEET_NAME]
    ws_tracker = wb[TRACKER_TABLE_SHEET_NAME]

    response = {
        "date": date,
        "priority_tasks": {},
        "reminders": [],
        "score": {
            "score_id": None,
            "threshold": 0,
            "current_score": 0,
            "rewards": "",
            "punishment": ""
        },
        "tracker": None
    }

    # --------------------------------------------------
    # PRIORITY TASKS
    # --------------------------------------------------
    priority_table = get_table(ws, PRIORITY_TABLE)
    priority_col_map = get_table_column_index_map(ws, priority_table)
    priority_rows = get_table_rows(ws, priority_table, priority_col_map)

    priority_result = defaultdict(list)

    for row in priority_rows:
        cell_value = row[PRIORITY_COLUMNS["datetime"]]
        if not cell_value:
            continue

        if str(cell_value)[:10] == date:
            priority_result[row[PRIORITY_COLUMNS["task_type"]]].append({
                "task_id": row[PRIORITY_COLUMNS["task_id"]],
                "task_name": row[PRIORITY_COLUMNS["task_name"]],
                "status": row[PRIORITY_COLUMNS["status"]],
            })

    response["priority_tasks"] = priority_result

    # --------------------------------------------------
    # REMINDERS
    # --------------------------------------------------
    reminder_table = get_table(ws_reminder, REMINDER_TABLE)
    reminder_col_map = get_table_column_index_map(ws_reminder, reminder_table)
    reminder_rows = get_table_rows(ws_reminder, reminder_table, reminder_col_map)

    for row in reminder_rows:
        cell_value = row[REMINDER_COLUMNS["datetime"]]
        if not cell_value:
            continue

        if str(cell_value)[:10] == date:
            response["reminders"].append({
                "reminder_id": row[REMINDER_COLUMNS["reminder_id"]],
                "reminder": row[REMINDER_COLUMNS["reminder"]],
            })

    # --------------------------------------------------
    # SCORE
    # --------------------------------------------------
    
    score_table = get_table(ws_score, SCORE_TABLE)
    score_col_map = get_table_column_index_map(ws_score, score_table)
    score_rows = get_table_rows(ws_score, score_table, score_col_map)

    for row in score_rows:
        cell_value = row[SCORE_COLUMNS["date"]]
        if not cell_value:
            continue

        if str(cell_value)[:10] == date:
            response["score"] = {
                "score_id": row[SCORE_COLUMNS["score_id"]],
                "threshold": row[SCORE_COLUMNS["threshold"]],
                "current_score": row[SCORE_COLUMNS["score"]],
                "rewards": row[SCORE_COLUMNS["rewards"]],
                "punishment": row[SCORE_COLUMNS["punishment"]],
            }
            break

    # --------------------------------------------------
    # TRACKER
    # --------------------------------------------------
    tracker_col_map = get_tracker_column_index_map()
    datetime_col = tracker_col_map[TRACKER_COLUMNS["datetime"]]

    for r in range(2, ws_tracker.max_row + 1):
        cell_value = ws_tracker.cell(row=r, column=datetime_col).value
        if not cell_value:
            continue

        if str(cell_value)[:10] == date:
            tracker_data = {
                "tracker_id": ws_tracker.cell(
                    row=r,
                    column=tracker_col_map[TRACKER_COLUMNS["tracker_id"]]
                ).value,
                "datetime": cell_value,
                "values": {}
            }

            for key, header in TRACKER_COLUMNS.items():
                if key in ["datetime", "tracker_id"]:
                    continue

                tracker_data["values"][key] = ws_tracker.cell(
                    row=r,
                    column=tracker_col_map[header]
                ).value

            response["tracker"] = tracker_data
            break

    return response



# @app.post("/save-priority-tasks")
# --------------------------------------------------
# Save PRIORITY TASKS 
# --------------------------------------------------
def save_priority_tasks_service(payload):

    wb = load_workbook(EXCEL_FILE)
    ws = wb[DAILY_TASK_SHEET_NAME]
    ws_score = wb[SCORE_TABLE_SHEET_NAME]

    # ---------------- PRIORITY ----------------
    table = get_table(ws, PRIORITY_TABLE)
    col_map = get_table_column_index_map(ws, table)
    rows = get_table_rows(ws, table, col_map)

    ist_datetime = get_ist_datetime()
    selected_date = payload.selected_date

    incoming_ids = {t.task_id for t in payload.tasks}

    # ---------------- DELETE removed tasks ----------------
    for row in reversed(rows):
        row_date = str(row[PRIORITY_COLUMNS["datetime"]])[:10]

        if (
            row_date == selected_date
            and row[PRIORITY_COLUMNS["task_type"]] == payload.task_type
            and row[PRIORITY_COLUMNS["task_id"]] not in incoming_ids
        ):
            ws.delete_rows(row["_row"])  # ✅ shifts rows upward

    # ✅ IMPORTANT: reload rows after deletion
    rows = get_table_rows(ws, table, col_map)

    # ---------------- BUILD FAST LOOKUP (OPTIMIZATION) ----------------
    existing_map = {}
    for row in rows:
        row_date = str(row[PRIORITY_COLUMNS["datetime"]])[:10]

        if (
            row_date == selected_date
            and row[PRIORITY_COLUMNS["task_type"]] == payload.task_type
        ):
            existing_map[row[PRIORITY_COLUMNS["task_id"]]] = row

    # ---------------- INSERT / UPDATE ----------------
    for task in payload.tasks:

        row = existing_map.get(task.task_id)

        # ---------------- UPDATE ----------------
        if row:
            existing_name = row[PRIORITY_COLUMNS["task_name"]]
            existing_status = row[PRIORITY_COLUMNS["status"]]

            # ✅ ONLY update if changed
            if (
                existing_name != task.task_name
                or existing_status != task.status
            ):
                ws.cell(
                    row=row["_row"],
                    column=col_map[PRIORITY_COLUMNS["datetime"]],
                    value=ist_datetime
                )
                ws.cell(
                    row=row["_row"],
                    column=col_map[PRIORITY_COLUMNS["task_name"]],
                    value=task.task_name
                )
                ws.cell(
                    row=row["_row"],
                    column=col_map[PRIORITY_COLUMNS["status"]],
                    value=task.status
                )

        # ---------------- INSERT (NO GAP GUARANTEED) ----------------
        else:
            # ✅ Find last non-empty row
            last_row = 0
            for r in range(ws.max_row, 0, -1):
                if any(ws.cell(row=r, column=c).value for c in col_map.values()):
                    last_row = r
                    break

            insert_row = last_row + 1  # ✅ always next row (no gaps)

            ws.cell(
                row=insert_row,
                column=col_map[PRIORITY_COLUMNS["datetime"]],
                value=get_ist_datetime()
            )
            ws.cell(
                row=insert_row,
                column=col_map[PRIORITY_COLUMNS["task_id"]],
                value=task.task_id
            )
            ws.cell(
                row=insert_row,
                column=col_map[PRIORITY_COLUMNS["task_type"]],
                value=payload.task_type
            )
            ws.cell(
                row=insert_row,
                column=col_map[PRIORITY_COLUMNS["task_name"]],
                value=task.task_name
            )
            ws.cell(
                row=insert_row,
                column=col_map[PRIORITY_COLUMNS["status"]],
                value=task.status
            )

    # ---------------- SCORE UPDATE ----------------
    score_table = get_table(ws_score, SCORE_TABLE)
    score_col_map = get_table_column_index_map(ws_score, score_table)
    score_rows = get_table_rows(ws_score, score_table, score_col_map)

    current_score = len(payload.tasks)

    upsert_score_for_today(
        ws=ws_score,
        score_table=score_table,
        score_col_map=score_col_map,
        rows=score_rows,
        today_date=selected_date,
        current_score=current_score
    )

    wb.save(EXCEL_FILE)

    return {
        "message": "Priority tasks & score updated",
        "date": selected_date,
        "current_score": current_score
    }

# @app.post("/save-reminders")
# --------------------------------------------------
# Save REMINDER APIs (FIXED)
# --------------------------------------------------
def save_reminders_service(payload):
    """
    Save or update daily reminders in the Excel reminder table.

    This API performs the following operations:

    1. Loads the Excel workbook and accesses the daily task sheet.
    2. Reads the reminder table and existing reminder rows.
    3. Identifies reminders received from the frontend request.
    4. Deletes reminders that were removed from the frontend for the current day.
    5. Updates existing reminders if their reminder_id already exists.
    6. Inserts new reminders if they do not exist in the table.
    7. Saves the updated workbook.

    Parameters
    ----------
    payload : SaveReminderRequest
        Request body containing:
        - reminders : List of reminders with reminder_id and reminder text.

    Returns
    -------
    dict
        JSON response containing:
        - message : Confirmation message
        - count : Number of reminders processed
    """

    wb = load_workbook(EXCEL_FILE)
    ws_reminder = wb[REMINDER_TABLE_SHEET_NAME]

    table = get_table(ws_reminder, REMINDER_TABLE)
    col_map = get_table_column_index_map(ws_reminder, table)
    rows = get_table_rows(ws_reminder, table, col_map)

    ist_datetime = get_ist_datetime()
    today = ist_datetime[:10]

    incoming_ids = {r.reminder_id for r in payload.reminders}

    # DELETE removed reminders
    for row in reversed(rows):
        row_date = str(row[REMINDER_COLUMNS["datetime"]])[:10]

        if (
            row_date == today
            and row[REMINDER_COLUMNS["reminder_id"]] not in incoming_ids
        ):
            ws_reminder.delete_rows(row["_row"])

    # INSERT / UPDATE
    for reminder in payload.reminders:
        matched = False

        for row in rows:
            row_date = str(row[REMINDER_COLUMNS["datetime"]])[:10]

            if (
                row_date == today
                and row[REMINDER_COLUMNS["reminder_id"]] == reminder.reminder_id
            ):
                ws_reminder.cell(row=row["_row"], column=col_map[REMINDER_COLUMNS["datetime"]], value=ist_datetime)
                ws_reminder.cell(row=row["_row"], column=col_map[REMINDER_COLUMNS["reminder"]], value=reminder.reminder)
                matched = True
                break

        if not matched:
            insert_row_into_table(
                ws_reminder,
                table,
                col_map,
                {
                    REMINDER_COLUMNS["datetime"]: ist_datetime,
                    REMINDER_COLUMNS["reminder_id"]: reminder.reminder_id,
                    REMINDER_COLUMNS["reminder"]: reminder.reminder
                }
            )

    wb.save(EXCEL_FILE)

    return {
        "message": "Reminders saved successfully",
        "count": len(payload.reminders)
    }



# @app.post("/save-score")
# --------------------------------------------------
# Save SCORE APIs
# --------------------------------------------------
def save_score_service(payload):
    """
    Save or update the daily score record in the Excel score table.

    This API manages the daily scoring information used for tracking
    productivity or performance metrics.

    Workflow:
    1. Loads the Excel workbook and accesses the score table.
    2. Checks whether a score record already exists for the given date.
    3. If a record exists, updates the score details.
    4. If no record exists, inserts a new row into the score table.
    5. Expands the Excel table range if a new row is inserted.
    6. Saves the workbook.

    Parameters
    ----------
    payload : SaveScoreRequest
        Request body containing:
        - date : Date for which the score is recorded
        - score_id : Unique identifier for the score entry
        - threshold : Target score threshold
        - current_score : Actual score achieved
        - rewards : Reward description for achieving the threshold
        - punishment : Punishment description if threshold is not met

    Returns
    -------
    dict
        JSON response containing:
        - message : Confirmation message
        - date : Date of the score entry
        - updated : Boolean indicating whether the record was updated or newly created
    """

    wb = load_workbook(EXCEL_FILE)
    ws_score = wb[SCORE_TABLE_SHEET_NAME]

    table = get_table(ws_score, SCORE_TABLE)
    col_map = get_table_column_index_map(ws_score, table)
    rows = get_table_rows(ws_score, table, col_map)

    matched = False

    # ---------- UPDATE if date exists ----------
    for row in rows:
        cell_value = row[SCORE_COLUMNS["date"]]

        if not cell_value:
            continue

        row_date = str(cell_value)[:10]

        if row_date == payload.date:
            ws_score.cell(
                row=row["_row"],
                column=col_map[SCORE_COLUMNS["date"]],
                value=payload.date
            )
            ws_score.cell(
                row=row["_row"],
                column=col_map[SCORE_COLUMNS["score_id"]],
                value=payload.score_id
            )
            ws_score.cell(
                row=row["_row"],
                column=col_map[SCORE_COLUMNS["threshold"]],
                value=payload.threshold
            )
            ws_score.cell(
                row=row["_row"],
                column=col_map[SCORE_COLUMNS["score"]],
                value=payload.current_score
            )
            ws_score.cell(
                row=row["_row"],
                column=col_map[SCORE_COLUMNS["rewards"]],
                value=payload.rewards
            )
            ws_score.cell(
                row=row["_row"],
                column=col_map[SCORE_COLUMNS["punishment"]],
                value=payload.punishment
            )

            matched = True
            break

    # ---------- INSERT if date not found ----------
    if not matched:
        ws_score.append([
            payload.date,
            payload.score_id,
            payload.threshold,
            payload.current_score,
            payload.rewards,
            payload.punishment
        ])

        # expand table range
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        table.ref = (
            f"{ws_score.cell(min_row, min_col).coordinate}:"
            f"{ws_score.cell(ws_score.max_row, max_col).coordinate}"
        )

    wb.save(EXCEL_FILE)

    return {
        "message": "Score saved successfully",
        "date": payload.date,
        "updated": matched
    }


# @app.post("/save-tracker")
# --------------------------------------------------
# Tracker APIs
# --------------------------------------------------
def save_tracker_service(payload):
    """
    Save or update time tracker data in the Excel tracker table.

    This API records planned vs actual time values for different
    time slots throughout the day.

    Workflow:
    1. Loads the Excel workbook and accesses the tracker sheet.
    2. Retrieves the column mapping for tracker fields.
    3. Checks if a tracker entry already exists using tracker_id.
    4. If the tracker entry exists, updates the corresponding row.
    5. If the tracker entry does not exist, inserts a new row.
    6. Stores the timestamp of the update.
    7. Saves the workbook.

    Parameters
    ----------
    payload : SaveTrackerRequest
        Request body containing:
        - tracker_id : Unique identifier for the tracker record
        - values : Dictionary of time-slot values (planned vs actual)

    Returns
    -------
    dict
        JSON response containing:
        - message : Confirmation message
        - tracker_id : Identifier of the tracker record
        - updated : Boolean indicating whether the record was updated or newly created
    """

    wb = load_workbook(EXCEL_FILE)
    ws_tracker = wb[TRACKER_TABLE_SHEET_NAME]

    col_map = get_tracker_column_index_map()
    ist_datetime = get_ist_datetime()

    existing_row = find_tracker_row(ws_tracker, col_map, payload.tracker_id)

    # ---------------- UPDATE ----------------
    if existing_row:
        ws_tracker.cell(
            row=existing_row,
            column=col_map[TRACKER_COLUMNS["datetime"]],
            value=ist_datetime
        )

        for key, value in payload.values.items():
            if key in TRACKER_COLUMNS:
                ws_tracker.cell(
                    row=existing_row,
                    column=col_map[TRACKER_COLUMNS[key]],
                    value=value
                )

        wb.save(EXCEL_FILE)

        return {
            "message": "Tracker updated successfully",
            "tracker_id": payload.tracker_id,
            "updated": True
        }

    # ---------------- INSERT ----------------
    insert_row = ws_tracker.max_row + 1

    ws_tracker.cell(
        row=insert_row,
        column=col_map[TRACKER_COLUMNS["datetime"]],
        value=ist_datetime
    )

    ws_tracker.cell(
        row=insert_row,
        column=col_map[TRACKER_COLUMNS["tracker_id"]],
        value=payload.tracker_id
    )

    for key, value in payload.values.items():
        if key in TRACKER_COLUMNS:
            ws_tracker.cell(
                row=insert_row,
                column=col_map[TRACKER_COLUMNS[key]],
                value=value
            )

    wb.save(EXCEL_FILE)

    return {
        "message": "Tracker created successfully",
        "tracker_id": payload.tracker_id,
        "updated": False
    }

