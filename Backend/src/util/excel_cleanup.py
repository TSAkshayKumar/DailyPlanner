from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter


def remove_empty_rows_from_table(ws, table):
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    rows_to_delete = []

    # 🚨 IMPORTANT:
    # min_row       -> header
    # min_row + 1   -> FIRST data row (KEEP EVEN IF EMPTY)
    # min_row + 2+  -> allowed to delete if empty
    for r in range(min_row + 2, max_row + 1):
        if all(
            ws.cell(row=r, column=c).value in (None, "")
            for c in range(min_col, max_col + 1)
        ):
            rows_to_delete.append(r)

    # delete bottom-up
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    # 🔄 Recalculate table range
    new_max_row = ws.max_row
    if new_max_row >= min_row + 1:
        table.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{new_max_row}"
        )


def remove_empty_rows_from_all_tables(excel_file, sheet_names):
    wb = load_workbook(excel_file)

    cleaned_tables = {}

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        cleaned_tables[sheet_name] = []

        for table_name in ws.tables:
            table = ws.tables[table_name]  # Table object
            remove_empty_rows_from_table(ws, table)
            cleaned_tables[sheet_name].append(table_name)

    wb.save(excel_file)
    return cleaned_tables