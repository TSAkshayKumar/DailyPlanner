from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter, range_boundaries
import pytz
from util.constant import (
    IST_TIMEZONE,
    SCORE_COLUMNS,
    TRACKER_COLUMNS,
    TRACKER_START_COL,
    GOAL_TABLE
)
IST = pytz.timezone(IST_TIMEZONE)

# --------------------------------------------------
# DATETIME HELPERS
# --------------------------------------------------

def get_ist_datetime() -> str:
    """
    Returns current datetime in IST formatted string.
    Example: '2026-03-15 09:30:12'
    """
    return datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")


# --------------------------------------------------
# EXCEL TABLE HELPERS
# --------------------------------------------------

def get_table(ws, table_name):
    """
    Returns Excel table object from worksheet.
    """
    if table_name not in ws.tables:
        raise ValueError(f"Excel table '{table_name}' not found")

    return ws.tables[table_name]


def get_table_column_index_map(ws, table):
    """
    Creates mapping:
    {
        column_header : column_index
    }
    """

    min_col, min_row, max_col, _ = range_boundaries(table.ref)

    headers = [
        ws.cell(row=min_row, column=col).value
        for col in range(min_col, max_col + 1)
    ]

    return {
        header: idx + min_col
        for idx, header in enumerate(headers)
    }


def get_table_rows(ws, table, col_map):
    """
    Reads rows from Excel table and returns
    list of dictionaries.
    """

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    rows = []

    for r in range(min_row + 1, max_row + 1):

        row_data = {
            col: ws.cell(row=r, column=idx).value
            for col, idx in col_map.items()
        }

        row_data["_row"] = r
        rows.append(row_data)

    return rows


def insert_row_into_table(ws, table, col_map, values_dict):
    """
    Inserts new row at end of Excel table
    while preserving table boundaries.
    """

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    new_row = max_row + 1

    for column_name, value in values_dict.items():

        ws.cell(
            row=new_row,
            column=col_map[column_name],
            value=value
        )

    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_row}"
    )


# --------------------------------------------------
# SCORE HELPERS
# --------------------------------------------------

def upsert_score_for_today(
    ws,
    score_table,
    score_col_map,
    rows,
    today_date,
    current_score
):
    """
    If today's score exists → update
    else insert new score row.
    """

    # -------- UPDATE EXISTING --------
    for row in rows:

        row_date = str(row[SCORE_COLUMNS["date"]])[:10]

        if row_date == today_date:

            ws.cell(
                row=row["_row"],
                column=score_col_map[SCORE_COLUMNS["score"]],
                value=current_score
            )

            return

    # -------- INSERT NEW --------
    insert_row_into_table(
        ws,
        score_table,
        score_col_map,
        {
            SCORE_COLUMNS["date"]: today_date,
            SCORE_COLUMNS["score_id"]: f"S-{today_date}",
            SCORE_COLUMNS["threshold"]: 0,
            SCORE_COLUMNS["score"]: current_score,
            SCORE_COLUMNS["rewards"]: "",
            SCORE_COLUMNS["punishment"]: ""
        }
    )


# --------------------------------------------------
# TRACKER HELPERS
# --------------------------------------------------

def get_tracker_column_index_map():
    """
    Returns mapping of tracker column header → column index
    """

    col_map = {}

    current_col = TRACKER_START_COL

    for header in TRACKER_COLUMNS.values():

        col_map[header] = current_col
        current_col += 1

    return col_map


def find_tracker_row(ws, col_map, tracker_id):
    """
    Finds row index for tracker_id
    """

    tracker_id_col = col_map[TRACKER_COLUMNS["tracker_id"]]

    for row in range(2, ws.max_row + 1):

        if ws.cell(row=row, column=tracker_id_col).value == tracker_id:
            return row

    return None


# --------------------------------------------------
# GOAL HELPERS
# --------------------------------------------------

def get_goal_rows(ws):
    """
    Returns goal table, column map, and rows
    """

    table = get_table(ws, GOAL_TABLE)

    col_map = get_table_column_index_map(ws, table)

    rows = get_table_rows(ws, table, col_map)

    return table, col_map, rows


# --------------------------------------------------
# DATE RANGE HELPERS
# --------------------------------------------------

def get_last_three_month_window(month: int, year: int, tz):
    """
    Returns date range covering last 3 months window.

    start_date = first day of (current month - 2)
    end_date   = last day of current month
    """

    current_month_start = datetime(year, month, 1, tzinfo=tz)

    start_date = current_month_start - relativedelta(months=2)

    end_date = (
        current_month_start
        + relativedelta(months=1)
        - relativedelta(seconds=1)
    )

    return start_date, end_date


def get_last_7_days_range(current_date_str: str):
    """
    Returns start and end datetime for last 7 days.
    """

    IST = pytz.timezone(IST_TIMEZONE)

    current_date = datetime.strptime(current_date_str, "%Y-%m-%d")

    current_date = IST.localize(current_date)

    start_date = current_date - timedelta(days=6)

    start_date = start_date.replace(
        hour=0,
        minute=0,
        second=0
    )

    end_date = current_date.replace(
        hour=23,
        minute=59,
        second=59
    ) 

    return start_date, end_date